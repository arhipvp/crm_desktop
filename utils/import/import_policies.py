import datetime as dt
import os
import sys
import logging

from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication

from ...database.db import db
from ...database.models import Policy
from ...services.clients import get_or_create_client_by_name
from ...services.policies import add_policy
from ...ui.common.client_import_dialog import ClientImportDialog

EXCEL_FILENAME = "policies_import.xlsx"

logger = logging.getLogger(__name__)


def parse_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    try:
        return dt.datetime.strptime(str(value), "%d.%m.%Y").date()
    except Exception:
        return None


def run_import():
    filepath = os.path.join(os.path.dirname(__file__), EXCEL_FILENAME)
    wb = load_workbook(filename=filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    success, skipped = 0, 0
    # Убедимся, что есть QApplication
    if not QApplication.instance():
        app = QApplication(sys.argv)

    with db.atomic():
        for row in rows:
            data = dict(zip(headers, row))
            client_name = str(data.get("ФИО клиента", "")).strip()
            policy_number = str(data.get("Номер полиса", "")).strip()
            start_date = parse_date(data.get("Дата начала"))
            end_date = parse_date(data.get("Дата окончания"))

            if not client_name or not policy_number or not start_date or not end_date:
                logger.info("❌ Пропущено: не хватает обязательных данных → %s", data)
                skipped += 1
                continue

            # Создаём или получаем клиента
            client, created = get_or_create_client_by_name(client_name)

            if created or not client.phone:
                dlg = ClientImportDialog(
                    suggested_name=client.name, suggested_phone=client.phone
                )

                if dlg.exec():
                    client = dlg.client
                else:
                    logger.info(
                        "⛔ Импорт прерван по пользователю для клиента: %s", client_name
                    )
                    skipped += 1
                    continue

            # Проверка на дубликат номера
            original_number = policy_number
            suffix = 1
            while Policy.get_or_none(Policy.policy_number == policy_number):
                policy_number = f"{original_number}-{suffix}"
                suffix += 1

            # Добавляем полис
            # Добавляем полис через словарь
            policy_data = {
                "policy_number": policy_number,
                "client_id": client.id,
                "start_date": start_date,
                "end_date": end_date,
                "insurance_type": data.get("Тип страхования"),
                "insurance_company": data.get("Компания"),
                "contractor": data.get("Контрагент"),
                "sales_channel": data.get("Канал продаж"),
                "vehicle_brand": data.get("Марка авто"),
                "vehicle_model": data.get("Модель авто"),
                "vehicle_vin": data.get("VIN"),
                "note": data.get("Комментарий"),
            }

            policy = add_policy(
                **policy_data,
                payments=[
                    {
                        "amount": float(data.get("Сумма", 0)) or 0,
                        "payment_date": start_date,
                    }
                ],
                first_payment_paid=True,
            )

            logger.info(
                "✅ Добавлен полис: %s (клиент: %s)", policy.policy_number, client.name
            )
            success += 1

    logger.info("\n✅ Импорт завершён: %s добавлено, %s пропущено.", success, skipped)


if __name__ == "__main__":
    run_import()
